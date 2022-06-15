#importação de bibliotecas utilizadas
import pytesseract as ocr
import numpy as np
import cv2
from PIL import Image, ImageChops
import time
import openpyxl

#inicia cronômetro para determinar tempo total de execução
start_time = time.time()

#acessar o caminho local do path tesseract
ocr.pytesseract.tesseract_cmd = r'D:\Usuários\notebook_william_lenovo\OneDrive\Programação\zz_PATH e Arquivos\Tesseract-OCR\tesseract.exe'

#função para cortar a área da foto de coleta da imagem
def coletar_area(img, area):
    
    #usando pillow (PIL)
    crop_img = img.crop(area)
    crop_img = ImageChops.invert(crop_img)
         
    return crop_img

#função para tratar imagem cortada antes de usar o tesseract para converter imagem em texto
def tratamento_imagem(img, sc=1.2):
    
    #convertendo imagem em um array editável de numpy[x, y, CANALS]
    npimagem = np.asarray(img).astype(np.uint8)  
    
    #redimensionando (zoom) foto através do sc definido
    img_zoom = cv2.resize(npimagem, None, fx=sc, fy=sc, interpolation=cv2.INTER_CUBIC)
    
    #desfoque gaussiano em imagem
    img_zoom = cv2.GaussianBlur(img_zoom, (5, 5), 0)
    
    # diminuição dos ruidos antes da binarização da imagem
    img_zoom[:, :, 0] = 0 # zerando o canal R (RED)
    img_zoom[:, :, 2] = 0 # zerando o canal B (BLUE)
    
    # atribuição em escala de cinza na imagem
    im = cv2.cvtColor(img_zoom, cv2.COLOR_RGB2GRAY)
        
    # aplicação da truncagem binária para a intensidade
    # pixels de intensidade de cor abaixo de 127 serão convertidos para 0 (PRETO)
    # pixels de intensidade de cor acima de 127 serão convertidos para 255 (BRANCO)
    # A atribuição do THRESH_OTSU incrementa uma análise inteligente dos nivels de truncagem
    ret, thresh = cv2.threshold(im, 127, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU) 
    
    # reconvertendo com o retorno do threshold em um objeto do tipo PIL.Image
    binimagem = Image.fromarray(thresh) 
  
    #retorna a imagem tratada
    return binimagem
'''
def inverter_cor (img):
    inv_img = ImageChops.invert(img)
    
    return inv_img
'''
#posição em pixel de cada coleta de dados, definido em tupla

#Geral
data = (1798, 8, 1914, 28)
hora = (1807, 29, 1907, 51)
nivel_canal = (70, 556, 121, 577)
temperatura = (1818, 58, 1904, 87)

#para UG1_fixo
UG1_FaseR_tensao = (10, 146, 95, 163)
UG1_FaseR_corrente = (108, 146, 192, 163)
i = 20 #um campo é 17 pixel abaixo do outro
UG1_potencia_ativa = (10, 146+i, 95, 163+i)
UG1_potencia_reativa = (108, 146+i, 192, 163+i)
UG1_fator_potencia = (10, 146+i*2, 95, 163+i*2)
UG1_distribuidor = (108, 146+i*2, 192, 163+i*2)
UG1_campo_excitacao_tensao = (10, 146+i*3, 95, 163+i*3)
UG1_campo_excitacao_corrente = (108, 146+i*3, 192, 163+i*3)

fixo = 139 #valor em y
#para UG2_fixo
UG2_FaseR_tensao = (10, 146+fixo, 95, 163+fixo)
UG2_FaseR_corrente = (108, 146+fixo, 192, 163+fixo)
i = 20 #um campo é 17 pixel abaixo do outro
UG2_potencia_ativa = (10, 146+i+fixo, 95, 163+i+fixo)
UG2_potencia_reativa = (108, 146+i+fixo, 192, 163+i+fixo)
UG2_fator_potencia = (10, 146+i*2+fixo, 95, 163+i*2+fixo)
UG2_distribuidor = (108, 146+i*2+fixo, 192, 163+i*2+fixo)
UG2_campo_excitacao_tensao = (10, 146+i*3+fixo, 95, 163+i*3+fixo)
UG2_campo_excitacao_corrente = (108, 146+i*3+fixo, 192, 163+i*3+fixo)

#UG1_variavel
UG1_acumulado_energia = (787, 474, 1003, 506)
UG1_velocidade = (851, 423, 956, 445)
UG1_potencia_ativa_oper = (490, 499, 560, 524)
UG1_potencia_aparente = (846, 337, 953, 361)

#anterior = 865
variavel = 875
#UG2_variavel
UG2_acumulado_energia = (787+variavel, 474, 1003+variavel, 506)
UG2_velocidade = (851+variavel, 423, 956+variavel, 445)
UG2_potencia_ativa_oper = (490+variavel, 499, 560+variavel, 524)
UG2_potencia_aparente = (846+variavel, 337, 953+variavel, 361)

#lista de variveis para coleta
geral = [data, hora, nivel_canal, temperatura]

#-----xxx-----
UG1_fixo = [UG1_FaseR_tensao, UG1_FaseR_corrente, UG1_potencia_ativa, UG1_potencia_reativa, 
            UG1_fator_potencia, UG1_distribuidor, UG1_campo_excitacao_tensao, UG1_campo_excitacao_corrente]
UG1_variavel = [UG1_acumulado_energia, UG1_velocidade, UG1_potencia_ativa_oper, UG1_potencia_aparente]

#-----xxx-----
UG2_fixo = [UG2_FaseR_tensao, UG2_FaseR_corrente, UG2_potencia_ativa, UG2_potencia_reativa, 
            UG2_fator_potencia, UG2_distribuidor, UG2_campo_excitacao_tensao, UG2_campo_excitacao_corrente]
UG2_variavel = [UG2_acumulado_energia, UG2_velocidade, UG2_potencia_ativa_oper, UG2_potencia_aparente]


#Coleta de dados para string por OCR e armazenar em planilha  

#abrindo e ativando o a planilha de excel modelo
wb = openpyxl.load_workbook("model_bd_coleta.xlsx")
planilha = wb.active

print("Iniciando a leitura das imagens...")

#coletando para 2022-02 - lote - 1
#coletando para 2022-02 - lote - 2.1
#coletando para 2022-02 - lote - 2.2
#coletando para 2022-02 - lote - 3

#>>>>> coletando para 2022-01 - lote - 1 - 1 a 1700
#coletando para 2022-01 - lote - 1 - 1701 a 4638, erro em 2107

#l é o valor de linhas
for l in range(1700, 4638):
       
    try:
        #iniciando o cronometro para cada imagem
        tempo_foto = time.time()
        
        # tipando a leitura para os canais de ordem RGB
        local_img = f"img_coleta/2022-03/01/img ({l}).jpeg"
        l=l+1
        imagem = Image.open(local_img).convert("RGB")
          
        
        #inicio da Coluna e retornando ao inicio
        c=0
        
        #planilha.cell(row=l, column=c).value = "teste"
        
        #Geral
        #Coleta da Geral
        k = 0
        for k in range(len(geral)):
            
            area = geral[k]
            
            img_tratada = coletar_area(imagem, area)
            #img_tratada.show()
            
            img_tratada = tratamento_imagem(img_tratada, 4)
            #img_tratada.show()
        
            texto = ocr.image_to_string(img_tratada)#, timeout=4)
            
            c = c+1
            planilha.cell(row=l, column=c).value = texto[:-1]
            #print(texto[:-1])
               
            k+=k
        
        
        #Coleta da UG1_fixo
        k = 0
        for k in range(len(UG1_fixo)):
            
            area = UG1_fixo[k]
            
            img_tratada = coletar_area(imagem, area)
            #img_tratada.show()
            
            img_tratada = tratamento_imagem(img_tratada, 20)
            #img_tratada.show()
               
            texto = ocr.image_to_string(img_tratada)#, timeout=2)
            
            c = c+1
            planilha.cell(row=l, column=c).value = texto[:-1]
            #print(texto[:-1])
                  
            k+=k
        #Coleta da UG1_variavel
        k = 0
        for k in range(len(UG1_variavel)):
            
            area = UG1_variavel[k]
            
            img_tratada = coletar_area(imagem, area)
            #img_tratada.show()
            
            img_tratada = tratamento_imagem(img_tratada, 6)
            #img_tratada.show()
               
            texto = ocr.image_to_string(img_tratada)#, timeout=4)
            
            c = c+1
            planilha.cell(row=l, column=c).value = texto[:-1]
            #print(texto[:-1])
        
            k+=k
        
        #Coleta da UG2_fixo
        k = 0
        for k in range(len(UG2_fixo)):
            area = UG2_fixo[k]
            
            img_tratada = coletar_area(imagem, area)
            #img_tratada.show()
            
            img_tratada = tratamento_imagem(img_tratada, 20)
            #img_tratada.show()
            
            texto = ocr.image_to_string(img_tratada)#, timeout=2)
            
            c = c+1
            planilha.cell(row=l, column=c).value = texto[:-1]
            #print(texto[:-1])
            
            k+=k
        
        #Coleta da UG2_variavel
        k = 0
        for k in range(len(UG2_variavel)):
            
            area = UG2_variavel[k]
            
            img_tratada = coletar_area(imagem, area)
            #img_tratada.show()
            
            img_tratada = tratamento_imagem(img_tratada, 6)
            #img_tratada.show()
            
            texto = ocr.image_to_string(img_tratada)#, timeout=4)
            
            c = c+1
            planilha.cell(row=l, column=c).value = texto[:-1]
            #print(texto[:-1])
            
            k+=k
        
        #entrar no loop - salvar por foto
        wb.save("coleta_2022_03-lote_1.2.xlsx")
        
        
        print(f"Salvo Foto [{l-1}] | --- Tempo Foto: %.6s segundos ---" %(time.time() - tempo_foto))
        
    except:
        print(f"Error foto: img [{l-1}]")
        pass

#fechando a planilha de excel
wb.close()

#apresentando o Print para cada imagem coletado com o tempo para acompanhemento do tempo
print("--- Tempo Total: %.6s segundos ---" % (time.time() - start_time)) 